import os
import PyPDF2
import re
import locale
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def clean_text(text):
    cleaned_text = ''
    for char in text:
        if ord(char) < 128:  # Filtra caracteres ASCII estándar
            cleaned_text += char
    return cleaned_text

def procesar_factura_pdf(ruta_pdf):
    # Abre el archivo PDF
    with open(ruta_pdf, 'rb') as file:
        # Crea un objeto de lectura de PDF
        reader = PyPDF2.PdfReader(file)
        
        # Extrae el texto de todas las páginas
        text = ''
        for page in reader.pages:
            text += page.extract_text()

    # Limpia el texto de caracteres no reconocidos
    cleaned_text = clean_text(text)

    # Expresiones regulares para extraer información
    codigo_factura_regex = r'FACTURAB\(Cod\.(\d+)\)'
    fecha_regex = r'Fecha(\d+/\d+/\d+)Hora\d+:\d+:\d+'
    subtotal_regex = r'SUBTOTALSINDESCUENTOS\$(.*?)DESCUENTOS'
    total_regex = r'TOTAL\$(\d+\.\d+)'

    # Extraer información utilizando expresiones regulares
    codigo_factura_match = re.search(codigo_factura_regex, cleaned_text)
    fecha_match = re.search(fecha_regex, cleaned_text)
    subtotal_match = re.search(subtotal_regex, cleaned_text, re.DOTALL)
    total_match = re.search(total_regex, cleaned_text)

    # Guardar información en variables
    codigo_factura = codigo_factura_match.group(1) if codigo_factura_match else None
    fecha = fecha_match.group(1) if fecha_match else None
    subtotal_sin_descuento = subtotal_match.group(1).strip() if subtotal_match else None
    total = total_match.group(1) if total_match else None

    return codigo_factura, fecha, subtotal_sin_descuento, total

# Establecer el locale para el formato de moneda
locale.setlocale(locale.LC_ALL, '')

# Crear un nuevo libro de Excel y seleccionar la primera hoja
wb = Workbook()
ws = wb.active

# Escribir los encabezados
ws['A1'] = 'Código de Factura'
ws['B1'] = 'Fecha'
ws['C1'] = 'Subtotal sin Descuento'
ws['D1'] = 'Total'

# Obtener la lista de archivos en la carpeta "facturas"
carpeta_facturas = 'facturas'
archivos_facturas = os.listdir(carpeta_facturas)

# Inicializar el índice de la fila en la hoja de Excel
fila_actual = 2

# Lista para almacenar los subtotales
subtotales = []

# Iterar sobre cada archivo en la carpeta "facturas"
for archivo_factura in archivos_facturas:
    # Procesar la factura PDF
    ruta_pdf = os.path.join(carpeta_facturas, archivo_factura)
    codigo_factura, fecha, subtotal_sin_descuento, total = procesar_factura_pdf(ruta_pdf)

    # Escribir los datos en las celdas correspondientes
    ws['A' + str(fila_actual)] = codigo_factura
    ws['B' + str(fila_actual)] = fecha
    ws['C' + str(fila_actual)] = '$' + '{:,.0f}'.format(float(subtotal_sin_descuento.replace(',', '.'))) if subtotal_sin_descuento else None
    ws['D' + str(fila_actual)] = '$' + '{:,.0f}'.format(float(total.replace(',', '.'))) if total else None

    # Guardar el total
    if total:
        subtotales.append(float(total.replace(',', '.')))

    # Incrementar el índice de la fila
    fila_actual += 1

# Calcular el subtotal en Excel
columna_subtotal = get_column_letter(ws.max_column)  # Obtiene la letra de la última columna
ultima_fila = ws.max_row  # Obtiene el número de la última fila
ws['A{}'.format(ultima_fila + 1)] = 'Subtotal'
ws['D{}'.format(ultima_fila + 1)] = '$' + '{:,.0f}'.format(sum(subtotales))

# Guardar el libro de Excel
wb.save('informacion_facturas.xlsx')
